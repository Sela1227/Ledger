#!/usr/bin/env python3
"""
Excel → JSON 匯入腳本
將「總資產.xlsx」56 個月的歷史資料,轉成 ledger app 可匯入的 JSON 格式
自動處理帳戶名稱演化(富邦房貸 → 玉山房貸 → 玉山貸款合併)

使用方式:
    pip install openpyxl
    python3 import_excel.py 總資產__2025-09-01_.xlsx ledger-history.json

輸出後,在 app 設定頁點「匯入 JSON」即可載入歷史資料。
"""

import openpyxl
import json
import re
import sys
from datetime import datetime
from pathlib import Path

# ============================================================
# 帳戶名稱對應表(Excel 原名 → app 內 account ID)
# 處理跨年代的命名變化
# ============================================================
NAME_TO_ID = {
    # === 我的資產 ===
    'TD 美股': 'td_us',
    'TD美股': 'td_us',
    '永豐美股': 'sinopac_us',
    '虛擬幣': 'crypto',
    '台股': 'tw_stock',
    '永豐銀': 'sinopac_bank',
    '玉山銀': 'yushan_bank',
    '中信銀': 'ctbc_bank',
    '台新銀': 'taishin_bank',
    '王道銀': 'oway_bank',
    '普瑞默': 'primer',
    '兆豐南非幣': 'mega_zar',
    '勞退自提': 'labor_self',
    '勞退績效': 'labor_perf',
    '勞退績效(0324)': 'labor_perf',
    '房子': 'house',
    '土地': 'land',
    '雲象': 'yunxiang',         # 後來消失
    '現金': 'cash_legacy',       # 早期未拆分的現金欄位
    # === 我的負債 ===
    '欠饅頭': 'mantou',
    # 玉山貸款的多種寫法,全部歸到單一 ID
    '玉山貸款': 'yushan_loan',
    '玉山賃款': 'yushan_loan',   # 你的筆誤
    # 早期拆分的玉山房貸/信貸,合併計入 yushan_loan
    '玉山房貸 1': 'yushan_loan_part1',
    '玉山房貸 2': 'yushan_loan_part2',
    '玉山信貸 1': 'yushan_credit_part1',
    '玉山信貸 2': 'yushan_credit_part2',
    # 富邦房貸(2021 年中之前)
    '富邦房貸': 'fubon_loan',     # 同名兩筆需特別處理
    # 後期才出現
    '台新車貸': 'taishin_car_loan',
    # 其他
    '小彤借款': 'xt_loan',
    # === 代管:小孩 ===
    'Dora 股票': 'dora_stock',
    'Dora 現金': 'dora_cash',
    'Leo 股票': 'leo_stock',
    'Leo 現金': 'leo_cash',
    'Aaron 股票': 'aaron_stock',
    'Aaron 現金': 'aaron_cash',
    # === 代管:景翰 ===
    '景翰 股票': 'jh_stock',
    '景翰 借券': 'jh_loan',
    '景翰 現金': 'jh_cash',
}

# Account 結構合併規則:把多個 ID 加總成一個 ID
# 用於處理「玉山房貸/信貸拆四筆」→「玉山貸款一筆」
MERGE_INTO_YUSHAN_LOAN = [
    'yushan_loan_part1', 'yushan_loan_part2',
    'yushan_credit_part1', 'yushan_credit_part2',
]

# ============================================================
# 帳戶 metadata(供 app 完整 accounts 列表使用)
# ============================================================
ACCOUNT_DEFS = [
    # === 我的資產 ===
    ('td_us',         'TD 美股',     'stock',       'asset',     True,  'me'),
    ('sinopac_us',    '永豐美股',    'stock',       'asset',     True,  'me'),
    ('crypto',        '虛擬幣',      'crypto',      'asset',     True,  'me'),
    ('tw_stock',      '台股',        'stock',       'asset',     True,  'me'),
    ('sinopac_bank',  '永豐銀',      'cash',        'asset',     True,  'me'),
    ('yushan_bank',   '玉山銀',      'cash',        'asset',     True,  'me'),
    ('ctbc_bank',     '中信銀',      'cash',        'asset',     True,  'me'),
    ('taishin_bank',  '台新銀',      'cash',        'asset',     True,  'me'),
    ('oway_bank',     '王道銀',      'cash',        'asset',     True,  'me'),
    ('primer',        '普瑞默',      'other',       'asset',     False, 'me'),
    ('mega_zar',      '兆豐南非幣',  'cash',        'asset',     True,  'me'),
    ('labor_self',    '勞退自提',    'pension',     'asset',     False, 'me'),
    ('labor_perf',    '勞退績效',    'pension',     'asset',     False, 'me'),
    ('house',         '房子',        'real_estate', 'asset',     False, 'me'),
    ('land',          '土地',        'real_estate', 'asset',     False, 'me'),
    ('yunxiang',      '雲象',        'other',       'asset',     False, 'me'),
    ('cash_legacy',   '現金(舊)',   'cash',        'asset',     True,  'me'),
    # === 我的負債 ===
    ('yushan_loan',   '玉山貸款',    'loan',        'liability', True,  'me'),
    ('mantou',        '欠饅頭',      'loan',        'liability', True,  'me'),
    ('fubon_loan',    '富邦房貸',    'loan',        'liability', True,  'me'),
    ('taishin_car_loan', '台新車貸', 'loan',        'liability', True,  'me'),
    ('xt_loan',       '小彤借款',    'loan',        'liability', True,  'me'),
    # === 代管:小孩 ===
    ('dora_stock',    'Dora 股票',   'stock',       'asset',     True,  'dora'),
    ('dora_cash',     'Dora 現金',   'cash',        'asset',     True,  'dora'),
    ('leo_stock',     'Leo 股票',    'stock',       'asset',     True,  'leo'),
    ('leo_cash',      'Leo 現金',    'cash',        'asset',     True,  'leo'),
    ('aaron_stock',   'Aaron 股票',  'stock',       'asset',     True,  'aaron'),
    ('aaron_cash',    'Aaron 現金',  'cash',        'asset',     True,  'aaron'),
    # === 代管:景翰 ===
    ('jh_stock',      '景翰 股票',   'stock',       'asset',     True,  'jinghan'),
    ('jh_loan',       '景翰 借券',   'other',       'asset',     True,  'jinghan'),
    ('jh_cash',       '景翰 現金',   'cash',        'asset',     True,  'jinghan'),
]

# ============================================================
# 工具函式
# ============================================================
def parse_sheet_name_to_date(sheet_name):
    """
    分頁名稱 → ISO 日期 (YYYY-MM-DD)
    支援:'202509', '20221001', '202107' 等格式
    """
    s = str(sheet_name).strip()
    # 8 碼 (YYYYMMDD)
    if len(s) == 8 and s.isdigit():
        return f"{s[:4]}-{s[4:6]}-{s[6:8]}"
    # 6 碼 (YYYYMM)
    if len(s) == 6 and s.isdigit():
        return f"{s[:4]}-{s[4:6]}-01"
    return None


def extract_balances_from_sheet(ws):
    """
    從一張 sheet 抓出 (帳戶名稱 → 金額) 的對應
    處理同名重複(例如富邦房貸有兩筆),用 _2、_3 後綴區別
    """
    balances = {}
    name_count = {}
    for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
        name = row[0]
        val = row[1] if len(row) > 1 else None
        if not isinstance(name, str):
            continue
        name = name.strip()
        if not name or name in ('資產項目',):
            continue
        if val is None or not isinstance(val, (int, float)):
            continue

        # 處理同名重複(富邦房貸 × 2)
        if name in name_count:
            name_count[name] += 1
            unique_name = f"{name}__{name_count[name]}"
        else:
            name_count[name] = 1
            unique_name = name

        balances[unique_name] = float(val)
    return balances


def map_to_account_id(raw_name, dup_idx=1):
    """
    Excel 帳戶名 → app account ID
    處理特殊規則:富邦房貸有兩筆,用 dup_idx 區分
    """
    base_name = raw_name.split('__')[0]

    # 富邦房貸的特殊處理(同名兩筆)
    if base_name == '富邦房貸':
        if dup_idx == 1:
            return 'fubon_loan_1'
        else:
            return 'fubon_loan_2'

    return NAME_TO_ID.get(base_name)


def merge_balances(balances_by_raw_name):
    """
    將 raw 帳戶名稱對應到 app 的 account_id,並合併同 ID 的金額
    例如:玉山房貸 1 + 玉山房貸 2 + 玉山信貸 1 + 玉山信貸 2 → 全部加總到 yushan_loan

    回傳 {account_id: amount}
    """
    out = {}
    fubon_count = 0
    for raw_name, amt in balances_by_raw_name.items():
        base = raw_name.split('__')[0]
        dup = int(raw_name.split('__')[1]) if '__' in raw_name else 1
        if base == '富邦房貸':
            fubon_count += 1
            dup = fubon_count

        acc_id = map_to_account_id(base, dup)
        if acc_id is None:
            print(f"  ⚠ 未對應的帳戶: {raw_name} = {amt}")
            continue

        # 合併規則:把拆分的玉山房貸/信貸全部加進 yushan_loan
        if acc_id in MERGE_INTO_YUSHAN_LOAN:
            acc_id = 'yushan_loan'
        # 富邦房貸兩筆也合併
        if acc_id in ('fubon_loan_1', 'fubon_loan_2'):
            acc_id = 'fubon_loan'

        out[acc_id] = out.get(acc_id, 0) + amt

    return out


def extract_shared_block_from_sheet(ws):
    """
    從 sheet 抓「一菜雙石」(Dora/Leo/Aaron) 的資料
    位於 C-G 欄,從 12 列左右開始
    """
    out = {}
    members = {'Dora': 'dora', 'Leo': 'leo', 'Aaron': 'aaron'}

    # 掃描所有列,找 C 欄是 Dora/Leo/Aaron 的位置
    for row in ws.iter_rows(min_col=3, max_col=6, values_only=True):
        c_val = row[0]
        if not isinstance(c_val, str):
            continue
        c_val = c_val.strip()
        if c_val in members:
            stock = row[1] if isinstance(row[1], (int, float)) else None
            cash = row[2] if isinstance(row[2], (int, float)) else None
            prefix = members[c_val]
            if stock is not None:
                out[f'{prefix}_stock'] = float(stock)
            if cash is not None:
                out[f'{prefix}_cash'] = float(cash)

    return out


def extract_jinghan_from_sheet(ws):
    """
    景翰的資料在最後一列(通常 row 22-24),
    格式:景翰 | 股票 | 借券 | 現金 | 總和 | 上月 | 差額
    """
    out = {}
    for row in ws.iter_rows(min_col=3, max_col=6, values_only=True):
        c_val = row[0]
        if isinstance(c_val, str) and c_val.strip() == '景翰':
            stock = row[1] if isinstance(row[1], (int, float)) else None
            loan = row[2] if isinstance(row[2], (int, float)) else None  # 借券
            cash = row[3] if isinstance(row[3], (int, float)) else None
            if stock is not None:
                out['jh_stock'] = float(stock)
            if loan is not None:
                out['jh_loan'] = float(loan)
            if cash is not None:
                out['jh_cash'] = float(cash)
            break
    return out


def extract_yunxiang_and_qyld(ws):
    """
    從 C21 抓雲象股價(2025/07 之後位置改變,A 欄不再有)
    從 D19 抓 QYLD 年化報酬率(每月實際值,非固定)
    從 F4 抓「一菜雙石+儲備」 → monthly_kid_wan(F4 是逐月變動的)
    從 I2 抓特斯拉車貸月繳 → tesla_pmt_wan(I2 是逐月變動,2025/11 起出現)
    """
    out = {}
    # C19 = 'QYLD 年化報酬率', D19 = 0.0739...
    # C21 = '雲象', D21 = 165.51
    for r in range(15, 25):
        c_label = ws.cell(row=r, column=3).value  # C 欄
        d_value = ws.cell(row=r, column=4).value  # D 欄
        if not isinstance(c_label, str): continue
        c_label = c_label.strip()
        if c_label == 'QYLD 年化報酬率' and isinstance(d_value, (int, float)):
            out['_qyld_yield'] = float(d_value)
        elif c_label == '雲象' and isinstance(d_value, (int, float)):
            out['yunxiang'] = float(d_value)

    # F4 = 一菜雙石+儲備(萬,逐月變動)— monthly_kid_wan
    f4 = ws['F4'].value
    if isinstance(f4, (int, float)) and f4 > 0:
        out['_monthly_kid_wan'] = float(f4)

    # I2 = 特斯拉車貸月繳(元,逐月變動;2025/11 後才出現)— tesla_pmt_wan(轉換成萬)
    i2 = ws['I2'].value
    if isinstance(i2, (int, float)) and i2 > 0:
        out['_tesla_pmt_wan'] = float(i2) / 10000.0  # 元 → 萬

    return out


# ============================================================
# 主流程
# ============================================================
def main():
    if len(sys.argv) < 2:
        print(f"用法: python3 {sys.argv[0]} <input.xlsx> [output.json]")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else Path('ledger-history.json')

    if not input_path.exists():
        print(f"找不到檔案: {input_path}")
        sys.exit(1)

    print(f"讀取: {input_path}")
    wb = openpyxl.load_workbook(input_path, data_only=True)

    snapshots = []
    for sheet_name in wb.sheetnames:
        date = parse_sheet_name_to_date(sheet_name)
        if not date:
            print(f"\n跳過(無法解析日期): {sheet_name}")
            continue

        print(f"\n處理 {sheet_name} → {date}")
        ws = wb[sheet_name]

        # A、B 欄:我的帳戶
        raw_balances = extract_balances_from_sheet(ws)
        my_balances = merge_balances(raw_balances)

        # C-F 欄:小孩帳戶
        kids_balances = extract_shared_block_from_sheet(ws)

        # 景翰
        jh_balances = extract_jinghan_from_sheet(ws)

        # 雲象 (C21) + 當月 QYLD 年化 (D19) + F4 一菜雙石 + I2 特斯拉
        yx_qyld = extract_yunxiang_and_qyld(ws)
        qyld = yx_qyld.pop('_qyld_yield', None)
        kid_wan = yx_qyld.pop('_monthly_kid_wan', None)
        tesla_pmt = yx_qyld.pop('_tesla_pmt_wan', None)

        # 合併
        all_balances = {**my_balances, **kids_balances, **jh_balances, **yx_qyld}

        # 過濾掉 0 值(早期欄位很多沒填)
        all_balances = {k: round(v, 4) for k, v in all_balances.items() if v != 0 or k in ('cash_legacy',)}

        snapshot = {
            'date': date,
            'balances': all_balances,
        }
        if qyld is not None:
            snapshot['qyld_yield'] = round(qyld, 6)
        if kid_wan is not None:
            snapshot['monthly_kid_wan'] = round(kid_wan, 4)
        if tesla_pmt is not None:
            snapshot['tesla_pmt_wan'] = round(tesla_pmt, 4)

        snapshots.append(snapshot)
        extras = []
        if qyld: extras.append(f"qyld={qyld:.4f}")
        if kid_wan: extras.append(f"kid={kid_wan:.2f}")
        if tesla_pmt: extras.append(f"tesla={tesla_pmt:.4f}")
        extra_str = (" + " + ", ".join(extras)) if extras else ""
        print(f"  ✓ {len(all_balances)} 個帳戶{extra_str}")

    # 排序(時間升冪)
    snapshots.sort(key=lambda s: s['date'])

    # 組合完整 JSON
    accounts = [
        {
            'id': aid,
            'name': name,
            'category': cat,
            'side': side,
            'liquid': liquid,
            'owner': owner,
            'active': True,
        }
        for (aid, name, cat, side, liquid, owner) in ACCOUNT_DEFS
    ]

    # 玉山貸款加上自動攤還(用最後一筆當基準)
    last_yushan = None
    if snapshots:
        last_yushan = snapshots[-1]['balances'].get('yushan_loan')
    if last_yushan is not None:
        for a in accounts:
            if a['id'] == 'yushan_loan':
                a['auto'] = {
                    'type': 'amortize',
                    'principal_wan': abs(last_yushan),
                    'rate': 0.021,
                    'years_left': 28,
                    'last_balance_wan': last_yushan,
                }

    # 偵測停用帳戶(從未在資料中出現的)
    used_ids = set()
    for s in snapshots:
        used_ids.update(s['balances'].keys())
    for a in accounts:
        if a['id'] not in used_ids:
            a['active'] = False
            print(f"  停用未使用帳戶: {a['name']}")

    output_data = {
        'version': 2,
        'settings': {
            'retirement_target_wan': 8955,
            'qyld_yield': 0.0739,
            'monthly_expense_wan': 21.81,
        },
        'accounts': accounts,
        'snapshots': snapshots,
    }

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)

    print(f"\n{'='*50}")
    print(f"✓ 完成 {len(snapshots)} 筆快照,寫入 {output_path}")
    print(f"  日期範圍: {snapshots[0]['date']} ~ {snapshots[-1]['date']}")
    print(f"  帳戶總數: {len(accounts)}(包含已停用)")
    print(f"{'='*50}")
    print(f"\n下一步:")
    print(f"1. 開啟 ledger app")
    print(f"2. 切換到「設定」頁")
    print(f"3. 點「匯入 JSON」→ 選擇 {output_path}")
    print(f"4. 確認覆蓋,即可看到完整歷史")


if __name__ == '__main__':
    main()
