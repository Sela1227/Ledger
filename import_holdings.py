#!/usr/bin/env python3
"""
import_holdings.py — V0.9.0 持股 fund_flows 匯入

從一或多個 .xlsx 抓 fund_flows,輸出可貼進 index.html 的 JSON 段
或當作 merge patch 從 app 「匯入 JSON」載入。

預期 Excel 格式(每個 broker 一份 .xlsx,或一份多 sheet):
  欄位順序固定:A=日期, B=匯入 USD(正數=匯入,負數=匯出), C=當日匯率
  Row 1 = header,row 2 起是資料
  日期可以是 Excel 原生 date 或 YYYY-MM-DD 字串

用法:
    pip install openpyxl

    python3 import_holdings.py \\
      --schwab schwab.xlsx --schwab-sheet flows \\
      --sinopac sinopac.xlsx --sinopac-sheet flows \\
      --firstrade firstrade.xlsx --firstrade-sheet flows \\
      --output patch.json

    # 或印出 paste-friendly 格式直接貼進 index.html DEFAULT_DATA.holdings.fund_flows
    python3 import_holdings.py --schwab schwab.xlsx --paste

輸出(--output patch.json):
{
  "version": 2,
  "holdings": {
    "fund_flows": {
      "schwab":    [{"date": "2014-04-29", "amount_usd": 99000, "fxrate": 30.25}, ...],
      "firstrade": [...],
      "sinopac":   [...]
    }
  }
}
"""
import argparse, json, sys
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print('請先 `pip install openpyxl`', file=sys.stderr)
    sys.exit(1)


def parse_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%Y.%m.%d'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None


def read_flows(xlsx_path, sheet_name):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]
        print(f'  ⚠ 找不到 sheet「{sheet_name}」,改讀第一個 sheet「{ws.title}」', file=sys.stderr)

    flows = []
    skipped = 0
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        date_raw, amt_raw, fx_raw = row
        date = parse_date(date_raw)
        if date is None:
            if any(v is not None for v in (date_raw, amt_raw, fx_raw)):
                skipped += 1
            continue
        try:
            amt = float(amt_raw)
            fx = float(fx_raw)
        except (TypeError, ValueError):
            skipped += 1
            continue
        if amt == 0 or fx <= 0:
            skipped += 1
            continue
        flows.append({
            'date': date,
            'amount_usd': round(amt, 2),
            'fxrate': round(fx, 4),
        })

    flows.sort(key=lambda f: f['date'])
    if skipped > 0:
        print(f'  ⚠ 略過 {skipped} 筆無效列(空/格式不對)', file=sys.stderr)
    return flows


def main():
    ap = argparse.ArgumentParser(
        description='V0.9.0 fund_flows Excel 匯入',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    for b in ('schwab', 'firstrade', 'sinopac'):
        ap.add_argument(f'--{b}', metavar='XLSX', help=f'{b} 的 .xlsx 路徑')
        ap.add_argument(f'--{b}-sheet', default='flows', metavar='NAME',
                        help=f'{b} sheet 名(預設 flows)')
    ap.add_argument('--output', '-o', metavar='FILE',
                    help='輸出 patch JSON(沒給就 stdout)')
    ap.add_argument('--paste', action='store_true',
                    help='印出可直接貼進 DEFAULT_DATA.holdings.fund_flows 的格式')

    args = ap.parse_args()

    fund_flows = {}
    for b in ('schwab', 'firstrade', 'sinopac'):
        path = getattr(args, b)
        if not path:
            continue
        sheet = getattr(args, f'{b}_sheet')
        print(f'[{b}] 讀 {path}(sheet={sheet})', file=sys.stderr)
        flows = read_flows(path, sheet)
        fund_flows[b] = flows
        sum_usd = sum(f['amount_usd'] for f in flows)
        if flows:
            span = f'{flows[0]["date"]} ~ {flows[-1]["date"]}'
        else:
            span = '(無資料)'
        print(f'  → {len(flows)} 筆 · USD {sum_usd:+,.0f} · {span}', file=sys.stderr)

    if not fund_flows:
        print('\n錯誤:未指定任何 broker 的 .xlsx,無事可做。', file=sys.stderr)
        print('用法:python3 import_holdings.py --schwab schwab.xlsx [--sinopac sinopac.xlsx ...]',
              file=sys.stderr)
        sys.exit(1)

    if args.paste:
        # paste-friendly:可貼進 index.html DEFAULT_DATA.holdings.fund_flows = { ... }
        print('\n貼進 index.html 的 DEFAULT_DATA.holdings.fund_flows = {')
        for b, flows in fund_flows.items():
            print(f'      // === {b}({len(flows)} 筆)===')
            print(f'      "{b}": [')
            for f in flows:
                print(f'        {{ "date": "{f["date"]}", "amount_usd": {f["amount_usd"]}, "fxrate": {f["fxrate"]} }},')
            print('      ],')
        print('    };')
    else:
        patch = {'version': 2, 'holdings': {'fund_flows': fund_flows}}
        out = json.dumps(patch, ensure_ascii=False, indent=2)
        if args.output:
            with open(args.output, 'w', encoding='utf-8') as f:
                f.write(out)
            print(f'\n已寫入 {args.output} — 在 app 設定頁「匯入 JSON」載入,選「合併」', file=sys.stderr)
        else:
            print(out)


if __name__ == '__main__':
    main()
